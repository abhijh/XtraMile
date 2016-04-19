import openpyxl
import datetime
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from django.http import HttpResponse
from control.forms import ParticipantForm


def index(request):
	if request.user.is_authenticated():
		return redirect("/control/upload/")	
	else :
		return render(request, 'control/index.html')
@login_required(login_url='/control/')	
def upload(request):
	return render(request, "control/upload.html")
@login_required(login_url='/control/')	
def writeToDB(request):
	handleUploadedFile(request.FILES["data_sheet"])
	filename = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")+'.xlsx'
	workbook = openpyxl.load_workbook(filename = filename)
	sheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
	for i in range(2, sheet.max_row+1):
		if sheet.cell(row = i, column = 1).value is not None:
			participant = {
				"name" : sheet.cell(row = i, column = 1).value,
				"email" : sheet.cell(row = i, column = 2).value,
				"phone" : sheet.cell(row = i, column = 3).value,
				"alt_phone" : sheet.cell(row = i, column = 4).value, 
				"occupation" : sheet.cell(row = i, column = 5).value,
				"organisation" : sheet.cell(row = i, column = 6).value,
				"how_know" : sheet.cell(row = i, column = 7).value
			}
		data = ParticipantForm(participant)
		if data.is_valid():
			data.save()
	messages.add_message(request, messages.SUCCESS, 'Data entered successfully')
	return redirect("/control/upload/", request)
def verify(request):
	username = request.POST['username']
	password = request.POST['password']
	user = authenticate(username=username, password=password)
	if user is not None:
	    if user.is_active:
	    	login(request, user)
	    	return redirect("/control/upload/")	
	    else:
	        print("The password is valid, but the account has been disabled!")
	else:
	    print("The username and password were incorrect.")	
	    return redirect("/control/")	
def handleUploadedFile(f):
	filename = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")+'.xlsx'
	with open(filename, 'w') as destination:
		for chunk in f.chunks():
			destination.write(chunk)
def logout_view(request):
	logout(request)
	return redirect("/control/")